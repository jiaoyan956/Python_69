#用来读取测试数据
from openpyxl import load_workbook

def reg_data(file_name,sheet_name):
    wb=load_workbook(file_name)
    sheet=wb[sheet_name]
    new_a=[]#存储所有行的测试用例数据
    for i in range(2,sheet.max_row+1):
        a=[]#某一行测试用例数据
        for j in range(1,sheet.max_column):
            a.append(sheet.cell(row=i,column=j).value)
        new_a.append(a)
    return new_a#返回所有测试用例数据

def write_data(file_name,sheet_name,row,column,value):#此函数是写入结果抖Excel中
    # 开始写入结果
    wb = load_workbook(file_name)
    sheet = wb[sheet_name]
    # 定位单元格存值   行  列   值
    sheet.cell(row=row, column=column).value =value
    #sheet.cell(row=rec_new_data[0] + 1, column=8).value = str(rec_response)
    # 进行判断，期望值与实际值是否相等，判断这个用例是否执行通过
    # actual = {'code': rec_response['code'], 'msg': rec_response['msg']}
    # if eval(rec_new_data[6]) == actual:
    #     print('测试用例执行通过')
    #     sheet.cell(row=rec_new_data[0] + 1, column=9).value = 'PASS'
    # else:
    #     print('测试用例执行不通过')
    #     sheet.cell(row=rec_new_data[0] + 1, column=9).value = 'FAIL'

    # 保存工作簿
    wb.save(file_name)

#只有在当前文件下面才会点击运行才会执行代码
if __name__ == '__main__':
    new_a=reg_data('python_69.xlsx','name_update')
    print(new_a)